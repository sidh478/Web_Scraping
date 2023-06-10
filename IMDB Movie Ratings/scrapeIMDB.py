from bs4 import BeautifulSoup
import requests
import openpyxl

excel=openpyxl.Workbook() #creating Excel file
print(excel.sheetnames) #checking sheets available
sheet=excel.active #selecting active sheet
sheet.title='Top Rated Movies' #updating sheet name
print(excel.sheetnames) #printing sheetname
sheet.append(['Movie rank','Movie Name','Year of Release','IMDB Rating']) #loading headings to excel

try:
    source=requests.get('https://www.imdb.com/chart/top') #requesting for url
    source.raise_for_status() #checking status of request

    soup=BeautifulSoup(source.text,'html.parser') #source text through html parser and create beautiful object

    movies=soup.find('tbody',class_="lister-list").find_all('tr') #finding parser in class lister-list and constraining with tr tag
    
    for movie in movies:
        name=movie.find('td',class_='titleColumn').a.text #extract name of movie
        rank=movie.find('td',class_='titleColumn').get_text(strip=True).split('.')[0] #extract rank of movie
        year=movie.find('td',class_='titleColumn').span.text.strip('()') #extract year of movie
        rating=movie.find('td',class_='ratingColumn imdbRating').strong.text #movie ratings
        print(rank,name,year,rating) #print list each movie details
        sheet.append([rank,name,year,rating]) #Adding same list to excel file

except Exception as e:
    print(e)

excel.save('IMDB Movie Ratings.xlsx')



